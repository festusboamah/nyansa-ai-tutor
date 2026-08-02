import csv
import io
import re
from datetime import date, datetime

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from academics.models import ClassEnrollment

from .models import SchoolMembership, StudentProfile


REQUIRED_HEADERS = {"student_id", "first_name", "last_name"}
OPTIONAL_HEADERS = {"gender", "date_of_birth", "guardian_name", "guardian_phone"}


def _normalise_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _parse_date(value, row_number, errors):
    if value in (None, ""):
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    errors.append(f"Row {row_number}: date_of_birth must use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY.")
    return ""


def _read_rows(uploaded_file):
    if uploaded_file.name.lower().endswith(".csv"):
        text = uploaded_file.read().decode("utf-8-sig")
        return list(csv.reader(io.StringIO(text)))
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def parse_student_roster(uploaded_file):
    raw_rows = _read_rows(uploaded_file)
    if not raw_rows:
        raise ValidationError("The roster file is empty.")
    headers = [_normalise_header(value) for value in raw_rows[0]]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValidationError(f"Missing required column(s): {', '.join(sorted(missing))}.")
    allowed = REQUIRED_HEADERS | OPTIONAL_HEADERS
    errors, records, seen_ids = [], [], set()
    for row_number, values in enumerate(raw_rows[1:], start=2):
        data = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers) if header in allowed}
        if not any(str(value or "").strip() for value in data.values()):
            continue
        record = {key: str(data.get(key) or "").strip() for key in allowed}
        record["student_id"] = record["student_id"].upper()
        if not record["student_id"]:
            errors.append(f"Row {row_number}: student_id is required.")
        elif record["student_id"] in seen_ids:
            errors.append(f"Row {row_number}: duplicate student_id {record['student_id']} in this file.")
        seen_ids.add(record["student_id"])
        if not record["first_name"] or not record["last_name"]:
            errors.append(f"Row {row_number}: first_name and last_name are required.")
        gender = record["gender"].upper()
        gender_map = {"F": "FEMALE", "FEMALE": "FEMALE", "M": "MALE", "MALE": "MALE", "OTHER": "OTHER"}
        if gender and gender not in gender_map:
            errors.append(f"Row {row_number}: gender must be Female, Male, or Other.")
        record["gender"] = gender_map.get(gender, "")
        record["date_of_birth"] = _parse_date(data.get("date_of_birth"), row_number, errors)
        records.append(record)
    if not records:
        errors.append("The file contains no student rows.")
    return records, errors


def _unique_username(school, student_id):
    User = get_user_model()
    base = re.sub(r"[^a-z0-9]+", "-", f"{school.slug}-{student_id.lower()}").strip("-")[:140]
    username, suffix = base, 1
    while User.objects.filter(username=username).exists():
        suffix += 1
        username = f"{base[:145]}-{suffix}"
    return username


@transaction.atomic
def import_student_roster(*, school_class, records):
    User = get_user_model()
    school = school_class.school
    created = updated = 0
    for record in records:
        membership = SchoolMembership.objects.filter(
            school=school, identifier=record["student_id"], role=SchoolMembership.Role.STUDENT
        ).select_related("user").first()
        if membership:
            user = membership.user
            updated += 1
        else:
            user = User(username=_unique_username(school, record["student_id"]), role=User.Role.STUDENT)
            user.set_unusable_password()
            user.save()
            membership = SchoolMembership.objects.create(
                school=school,
                user=user,
                role=SchoolMembership.Role.STUDENT,
                identifier=record["student_id"],
                portal_access_enabled=False,
            )
            created += 1
        user.first_name, user.last_name = record["first_name"], record["last_name"]
        user.save(update_fields=["first_name", "last_name"])
        profile, _ = StudentProfile.objects.update_or_create(
            membership=membership,
            defaults={
                "gender": record.get("gender", ""),
                "date_of_birth": record.get("date_of_birth") or None,
                "guardian_name": record.get("guardian_name", ""),
                "guardian_phone": record.get("guardian_phone", ""),
            },
        )
        profile.full_clean()
        ClassEnrollment.objects.get_or_create(school_class=school_class, student=membership)
    return created, updated
