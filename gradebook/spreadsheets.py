from decimal import Decimal, InvalidOperation
from io import BytesIO
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation


INSTRUCTIONS_SHEET = "Instructions"
ROSTER_SHEET = "Grade Roster"
ROSTER_HEADERS = ["Student Record ID", "Student Identifier", "Student Name", "Score"]


class WorkbookValidationError(ValueError):
    pass


def build_grade_template(*, assessment, enrollments):
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = INSTRUCTIONS_SHEET
    roster = workbook.create_sheet(ROSTER_SHEET)
    workbook.properties.title = f"Nyansa grade roster - {assessment.title}"
    workbook.properties.creator = "Nyansa"

    instructions.append(["Nyansa Grade Import Template"])
    instructions.append(["Assessment Record ID", assessment.pk])
    instructions.append(["School Record ID", assessment.school_id])
    instructions.append(["Assessment", assessment.title])
    instructions.append(["Class", assessment.offering.school_class.name])
    instructions.append(["Subject", assessment.offering.subject.name])
    instructions.append(["Maximum Score", float(assessment.max_score)])
    instructions.append([])
    instructions.append(["Instructions"])
    instructions.append(["Enter scores only in the blue Score column. Leave a score blank to skip that student."])
    instructions.append(["Do not add, remove, reorder, or replace students. Upload this file to the same assessment."])
    instructions.merge_cells("A1:D1")
    instructions["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="123B6D")
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 30
    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 24
    instructions.column_dimensions["C"].width = 18
    instructions.column_dimensions["D"].width = 18
    instructions["A9"].font = Font(bold=True, color="123B6D")
    instructions["A10"].alignment = Alignment(wrap_text=True)
    instructions["A11"].alignment = Alignment(wrap_text=True)
    instructions.row_dimensions[10].height = 32
    instructions.row_dimensions[11].height = 32
    instructions.sheet_view.showGridLines = False

    roster.append(ROSTER_HEADERS)
    for cell in roster[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123B6D")
        cell.alignment = Alignment(vertical="center")
    roster.row_dimensions[1].height = 26
    for row_number, enrollment in enumerate(enrollments, start=2):
        membership = enrollment.student
        name = membership.user.get_full_name() or membership.user.username
        roster.append([membership.pk, membership.identifier, name, None])
        roster.cell(row_number, 4).fill = PatternFill("solid", fgColor="DCEEFF")
        roster.cell(row_number, 4).protection = Protection(locked=False)
    roster.freeze_panes = "A2"
    roster.auto_filter.ref = f"A1:D{max(roster.max_row, 2)}"
    roster.column_dimensions["A"].width = 20
    roster.column_dimensions["B"].width = 22
    roster.column_dimensions["C"].width = 34
    roster.column_dimensions["D"].width = 16
    if roster.max_row >= 2:
        validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2=str(assessment.max_score),
            allow_blank=True,
        )
        validation.error = f"Enter a score from 0 to {assessment.max_score}."
        validation.errorTitle = "Invalid score"
        validation.prompt = f"Maximum score: {assessment.max_score}"
        validation.promptTitle = assessment.title
        validation.showErrorMessage = True
        validation.showInputMessage = True
        roster.add_data_validation(validation)
        validation.add(f"D2:D{roster.max_row}")
    roster.protection.sheet = True
    roster.protection.password = "nyansa"
    roster.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_grade_template(*, file, assessment, enrollments):
    try:
        with ZipFile(file) as archive:
            if len(archive.infolist()) > 100 or sum(item.file_size for item in archive.infolist()) > 10 * 1024 * 1024:
                raise WorkbookValidationError("The workbook is too large or complex to process safely.")
    except BadZipFile as error:
        raise WorkbookValidationError("The file is not a readable .xlsx workbook.") from error
    file.seek(0)
    try:
        workbook = load_workbook(file, read_only=True, data_only=False)
    except Exception as error:
        raise WorkbookValidationError("The file is not a readable .xlsx workbook.") from error
    if INSTRUCTIONS_SHEET not in workbook.sheetnames or ROSTER_SHEET not in workbook.sheetnames:
        raise WorkbookValidationError("This is not a Nyansa grade roster template.")
    instructions = workbook[INSTRUCTIONS_SHEET]
    if str(instructions["B2"].value) != str(assessment.pk) or str(instructions["B3"].value) != str(assessment.school_id):
        raise WorkbookValidationError("This workbook belongs to a different school or assessment.")
    roster = workbook[ROSTER_SHEET]
    headers = [roster.cell(1, column).value for column in range(1, 5)]
    if headers != ROSTER_HEADERS:
        raise WorkbookValidationError("The grade roster columns were changed. Download a fresh template.")

    enrolled = {row.student_id: row.student for row in enrollments}
    seen = set()
    parsed = []
    max_row = max(roster.max_row, 1)
    if max_row > len(enrolled) + 1:
        raise WorkbookValidationError("The workbook contains extra roster rows. Download a fresh template.")
    for row_number in range(2, max_row + 1):
        student_value = roster.cell(row_number, 1).value
        score_cell = roster.cell(row_number, 4)
        raw_student_id = "" if student_value is None else str(student_value).strip()
        raw_score = "" if score_cell.value is None else str(score_cell.value).strip()
        student = None
        error = ""
        try:
            numeric_student_id = Decimal(str(student_value))
            if not numeric_student_id.is_finite() or numeric_student_id != numeric_student_id.to_integral_value():
                raise ValueError
            student_id = int(numeric_student_id)
        except (InvalidOperation, TypeError, ValueError):
            student_id = None
            error = "Student record ID is missing or invalid."
        if student_id is not None:
            student = enrolled.get(student_id)
            if student is None:
                error = "Student is not actively enrolled in this assessment class."
            elif student_id in seen:
                error = "Student appears more than once in the workbook."
                student = None
            seen.add(student_id)

        score = None
        status = "ERROR" if error else "SKIPPED"
        if not error and raw_score:
            if score_cell.data_type == "f":
                error = "Formulas are not allowed in score cells."
            else:
                try:
                    score = Decimal(raw_score)
                    if not score.is_finite():
                        raise InvalidOperation
                    if score < 0 or score > assessment.max_score:
                        error = f"Score must be from 0 to {assessment.max_score}."
                except (InvalidOperation, ValueError):
                    error = "Enter a valid finite number."
            status = "ERROR" if error else "VALID"
        parsed.append({
            "row_number": row_number,
            "student": student,
            "raw_student_id": raw_student_id,
            "raw_score": raw_score,
            "score": score if not error else None,
            "status": status,
            "error": error,
        })

    missing = set(enrolled) - seen
    for student_id in sorted(missing):
        parsed.append({
            "row_number": len(parsed) + 2,
            "student": enrolled[student_id],
            "raw_student_id": str(student_id),
            "raw_score": "",
            "score": None,
            "status": "ERROR",
            "error": "Student row is missing from the workbook.",
        })
    workbook.close()
    return parsed
