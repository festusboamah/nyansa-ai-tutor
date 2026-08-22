from datetime import date
from decimal import Decimal
from io import BytesIO

from django.core.exceptions import ValidationError
from django.core.exceptions import PermissionDenied
from django.test import TestCase
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import models
from openpyxl import load_workbook

from academics.models import AcademicYear, ClassEnrollment, SchoolClass, SubjectOffering, TeacherAssignment, Term
from accounts.models import User
from courses.models import Subject
from schools.models import School, SchoolMembership

from quizzes.models import Assignment, AssignmentSubmission, Quiz, Submission
from . import evidence
from .history import record_grade_entry, review_grade_entry
from .models import Assessment, AssessmentCategory, GradeEntry, GradeEntryRevision, GradeImportBatch, GradeImportRow, GradeReviewDecision, GradeScheme
from .services import activate_grade_scheme, calculate_weighted_result, configure_ges_grade_scheme
from .sync import sync_legacy_assessment


class GradebookTests(TestCase):
    def setUp(self):
        self.school = School.objects.create(name="Nyansa School", slug="nyansa-school")
        self.year = AcademicYear.objects.create(
            school=self.school,
            name="2026/2027",
            start_date=date(2026, 9, 1),
            end_date=date(2027, 7, 31),
            is_current=True,
        )
        self.term = Term.objects.create(
            academic_year=self.year,
            name="Term 1",
            order=1,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 12, 15),
        )
        self.school_class = SchoolClass.objects.create(
            school=self.school, academic_year=self.year, name="JHS 1"
        )
        self.subject = Subject.objects.create(school=self.school, name="Mathematics")
        self.offering = SubjectOffering.objects.create(
            school=self.school,
            school_class=self.school_class,
            subject=self.subject,
            term=self.term,
        )
        self.student = self._membership("student", SchoolMembership.Role.STUDENT)
        self.teacher = self._membership("teacher", SchoolMembership.Role.TEACHER)
        ClassEnrollment.objects.create(school_class=self.school_class, student=self.student)
        self.scheme = GradeScheme.objects.create(
            school=self.school, academic_year=self.year, name="Standard"
        )
        self.coursework = AssessmentCategory.objects.create(
            scheme=self.scheme, name="Coursework", code="coursework", weight=Decimal("40.00"), order=1
        )
        self.exam = AssessmentCategory.objects.create(
            scheme=self.scheme, name="Exam", code="exam", weight=Decimal("60.00"), order=2
        )

    def _membership(self, username, role, school=None):
        user = User.objects.create_user(username=username, password="test-password")
        return SchoolMembership.objects.create(school=school or self.school, user=user, role=role)

    def _assessment(self, category, title, max_score="100.00"):
        return Assessment.objects.create(
            school=self.school,
            offering=self.offering,
            category=category,
            title=title,
            max_score=Decimal(max_score),
            status=Assessment.Status.PUBLISHED,
        )

    def test_scheme_activation_requires_weights_totaling_100_percent(self):
        self.exam.weight = Decimal("50.00")
        self.exam.save(update_fields=["weight"])
        with self.assertRaisesMessage(ValidationError, "current total is 90"):
            activate_grade_scheme(self.scheme)
        self.scheme.refresh_from_db()
        self.assertEqual(self.scheme.status, GradeScheme.Status.DRAFT)

    def test_configure_ges_scheme_preserves_and_archives_legacy_scheme(self):
        self.scheme.status = GradeScheme.Status.ACTIVE
        self.scheme.save(update_fields=["status"])
        configured = configure_ges_grade_scheme(self.year)
        self.scheme.refresh_from_db()
        self.assertEqual(self.scheme.status, GradeScheme.Status.ARCHIVED)
        self.assertEqual(configured.status, GradeScheme.Status.ACTIVE)
        self.assertEqual(
            list(configured.categories.values_list("name", "weight", "order")),
            [
                ("Class Score", Decimal("50.00"), 1),
                ("Examination", Decimal("50.00"), 2),
            ],
        )

    def test_school_admin_can_activate_ges_scheme_from_settings(self):
        admin = self._membership("grade-admin", SchoolMembership.Role.SCHOOL_ADMIN)
        self.client.force_login(admin.user)
        response = self.client.post(
            reverse("gradebook_settings"), {"academic_year": self.year.pk}, secure=True
        )
        self.assertRedirects(
            response,
            f"{reverse('gradebook_settings')}?academic_year={self.year.pk}",
            fetch_redirect_response=False,
        )
        self.assertTrue(GradeScheme.objects.filter(
            school=self.school,
            academic_year=self.year,
            name="GES 50/50 grading scheme",
            status=GradeScheme.Status.ACTIVE,
        ).exists())

    def test_teacher_cannot_open_grade_settings(self):
        self.client.force_login(self.teacher.user)
        response = self.client.get(reverse("gradebook_settings"), secure=True)
        self.assertRedirects(response, reverse("home"), fetch_redirect_response=False)

    def test_activating_scheme_archives_previous_scheme_for_year(self):
        old = GradeScheme.objects.create(
            school=self.school,
            academic_year=self.year,
            name="Old scheme",
            status=GradeScheme.Status.ACTIVE,
        )
        activate_grade_scheme(self.scheme)
        old.refresh_from_db()
        self.scheme.refresh_from_db()
        self.assertEqual(old.status, GradeScheme.Status.ARCHIVED)
        self.assertEqual(self.scheme.status, GradeScheme.Status.ACTIVE)

    def test_assessment_rejects_category_from_another_school(self):
        other = School.objects.create(name="Other School", slug="other-school")
        other_year = AcademicYear.objects.create(
            school=other,
            name="2026/2027",
            start_date=date(2026, 9, 1),
            end_date=date(2027, 7, 31),
        )
        other_scheme = GradeScheme.objects.create(school=other, academic_year=other_year, name="Other")
        other_category = AssessmentCategory.objects.create(
            scheme=other_scheme, name="Exam", code="exam", weight=Decimal("100"), order=1
        )
        assessment = Assessment(
            school=self.school,
            offering=self.offering,
            category=other_category,
            title="Wrong tenant",
            max_score=Decimal("100"),
        )
        with self.assertRaises(ValidationError):
            assessment.full_clean()

    def test_grade_rejects_score_above_maximum(self):
        assessment = self._assessment(self.coursework, "Quiz", "20")
        entry = GradeEntry(
            school=self.school,
            assessment=assessment,
            student=self.student,
            recorded_by=self.teacher,
            score=Decimal("21"),
        )
        with self.assertRaisesMessage(ValidationError, "Score cannot exceed"):
            entry.full_clean()

    def test_grade_rejects_student_not_enrolled_in_assessment_class(self):
        outsider = self._membership("outsider", SchoolMembership.Role.STUDENT)
        assessment = self._assessment(self.coursework, "Class test")
        entry = GradeEntry(
            school=self.school,
            assessment=assessment,
            student=outsider,
            recorded_by=self.teacher,
            score=Decimal("70"),
        )
        with self.assertRaisesMessage(ValidationError, "actively enrolled"):
            entry.full_clean()

    def test_grade_rejects_cross_school_recorder(self):
        other = School.objects.create(name="Other School", slug="other-school")
        recorder = self._membership("other-teacher", SchoolMembership.Role.TEACHER, school=other)
        assessment = self._assessment(self.coursework, "Project")
        entry = GradeEntry(
            school=self.school,
            assessment=assessment,
            student=self.student,
            recorded_by=recorder,
            score=Decimal("70"),
        )
        with self.assertRaisesMessage(ValidationError, "same school"):
            entry.full_clean()

    def test_weighted_result_uses_published_entries_and_reweights_available_categories(self):
        coursework = self._assessment(self.coursework, "Project", "50")
        exam = self._assessment(self.exam, "End-of-term exam")
        GradeEntry.objects.create(
            school=self.school,
            assessment=coursework,
            student=self.student,
            recorded_by=self.teacher,
            score=Decimal("40"),
            status=GradeEntry.Status.PUBLISHED,
        )
        GradeEntry.objects.create(
            school=self.school,
            assessment=exam,
            student=self.student,
            recorded_by=self.teacher,
            score=Decimal("50"),
            status=GradeEntry.Status.DRAFT,
        )
        result = calculate_weighted_result(
            student=self.student, offering=self.offering, scheme=self.scheme
        )
        self.assertEqual(result["final_score"], Decimal("80.00"))
        self.assertEqual(result["used_weight"], Decimal("40.00"))

    def test_weighted_result_combines_category_averages(self):
        coursework = self._assessment(self.coursework, "Project", "50")
        exam = self._assessment(self.exam, "Exam")
        for assessment, score in ((coursework, "40"), (exam, "50")):
            GradeEntry.objects.create(
                school=self.school,
                assessment=assessment,
                student=self.student,
                recorded_by=self.teacher,
                score=Decimal(score),
                status=GradeEntry.Status.PUBLISHED,
            )
        result = calculate_weighted_result(
            student=self.student, offering=self.offering, scheme=self.scheme
        )
        self.assertEqual(result["final_score"], Decimal("62.00"))
        self.assertEqual(result["used_weight"], Decimal("100.00"))

    def test_category_evidence_reports_approved_average_and_entry_count(self):
        coursework = self._assessment(self.coursework, "Project", "50")
        exam = self._assessment(self.exam, "Exam")
        GradeEntry.objects.create(
            school=self.school, assessment=coursework, student=self.student,
            recorded_by=self.teacher, score=Decimal("40"), status=GradeEntry.Status.PUBLISHED,
            review_status=GradeEntry.ReviewStatus.APPROVED,
        )
        GradeEntry.objects.create(
            school=self.school, assessment=exam, student=self.student,
            recorded_by=self.teacher, score=Decimal("30"), status=GradeEntry.Status.DRAFT,
        )
        rows = evidence.category_evidence(self.student, self.offering, self.scheme)
        by_name = {row["name"]: row for row in rows}
        self.assertEqual(by_name["Coursework"]["average"], Decimal("80"))
        self.assertEqual(by_name["Coursework"]["entry_count"], 1)
        self.assertIsNone(by_name["Exam"]["average"])
        self.assertEqual(by_name["Exam"]["entry_count"], 0)

    def test_active_scheme_for_returns_only_active_scheme_for_year(self):
        self.assertIsNone(evidence.active_scheme_for(self.school, self.year))
        self.scheme.status = GradeScheme.Status.ACTIVE
        self.scheme.save(update_fields=["status"])
        self.assertEqual(evidence.active_scheme_for(self.school, self.year), self.scheme)
        self.scheme.status = GradeScheme.Status.ARCHIVED
        self.scheme.save(update_fields=["status"])
        self.assertIsNone(evidence.active_scheme_for(self.school, self.year))


class TeacherGradebookWorkflowTests(TestCase):
    def setUp(self):
        self.school = School.objects.create(name="Workflow School", slug="workflow-school")
        self.year = AcademicYear.objects.create(
            school=self.school,
            name="2026/2027",
            start_date=date(2026, 9, 1),
            end_date=date(2027, 7, 31),
            is_current=True,
        )
        self.term = Term.objects.create(
            academic_year=self.year,
            name="Term 1",
            order=1,
            start_date=date(2026, 9, 1),
            end_date=date(2026, 12, 15),
        )
        self.school_class = SchoolClass.objects.create(
            school=self.school, academic_year=self.year, name="Basic 6"
        )
        self.subject = Subject.objects.create(school=self.school, name="English")
        self.offering = SubjectOffering.objects.create(
            school=self.school,
            school_class=self.school_class,
            subject=self.subject,
            term=self.term,
        )
        self.teacher_user = User.objects.create_user(
            username="grade-teacher", password="test-password", role=User.Role.TEACHER
        )
        self.teacher = SchoolMembership.objects.create(
            school=self.school, user=self.teacher_user, role=SchoolMembership.Role.TEACHER
        )
        TeacherAssignment.objects.create(offering=self.offering, teacher=self.teacher, is_lead=True)
        self.other_teacher_user = User.objects.create_user(
            username="other-grade-teacher", password="test-password", role=User.Role.TEACHER
        )
        SchoolMembership.objects.create(
            school=self.school, user=self.other_teacher_user, role=SchoolMembership.Role.TEACHER
        )
        self.student_memberships = []
        for number in (1, 2):
            user = User.objects.create_user(username=f"roster-student-{number}", password="test-password")
            membership = SchoolMembership.objects.create(
                school=self.school, user=user, role=SchoolMembership.Role.STUDENT
            )
            ClassEnrollment.objects.create(school_class=self.school_class, student=membership)
            self.student_memberships.append(membership)
        self.scheme = GradeScheme.objects.create(
            school=self.school,
            academic_year=self.year,
            name="Active scheme",
            status=GradeScheme.Status.ACTIVE,
        )
        self.category = AssessmentCategory.objects.create(
            scheme=self.scheme, name="Classwork", code="classwork", weight=Decimal("100"), order=1
        )
        self.assessment = Assessment.objects.create(
            school=self.school,
            offering=self.offering,
            category=self.category,
            title="Writing exercise",
            max_score=Decimal("20"),
            status=Assessment.Status.PUBLISHED,
        )

    def test_teacher_only_sees_assigned_offerings(self):
        other_subject = Subject.objects.create(school=self.school, name="Science")
        SubjectOffering.objects.create(
            school=self.school,
            school_class=self.school_class,
            subject=other_subject,
            term=self.term,
        )
        self.client.force_login(self.teacher_user)
        response = self.client.get(reverse("gradebook_offerings"), secure=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "English")
        self.assertNotContains(response, "Science")

    def test_unassigned_teacher_cannot_open_assessment_roster(self):
        self.client.force_login(self.other_teacher_user)
        response = self.client.get(reverse("gradebook_roster", args=[self.assessment.pk]), secure=True)
        self.assertEqual(response.status_code, 404)

    def test_student_membership_cannot_open_gradebook(self):
        self.client.force_login(self.student_memberships[0].user)
        response = self.client.get(reverse("gradebook_offerings"), secure=True)
        self.assertRedirects(response, reverse("home"), fetch_redirect_response=False)

    def test_teacher_can_create_assessment_from_active_scheme(self):
        self.client.force_login(self.teacher_user)
        response = self.client.post(
            reverse("gradebook_create_assessment", args=[self.offering.pk]),
            {
                "category": self.category.pk,
                "title": "Vocabulary quiz",
                "max_score": "15.00",
                "due_at": "",
                "status": Assessment.Status.DRAFT,
            },
            secure=True,
        )
        created = Assessment.objects.get(title="Vocabulary quiz")
        self.assertRedirects(response, reverse("gradebook_roster", args=[created.pk]), fetch_redirect_response=False)
        self.assertEqual(created.school, self.school)
        self.assertEqual(created.offering, self.offering)

    def test_assessment_form_provisions_default_categories_for_unconfigured_year(self):
        next_year = AcademicYear.objects.create(
            school=self.school,
            name="2027/2028",
            start_date=date(2027, 9, 1),
            end_date=date(2028, 7, 31),
        )
        next_term = Term.objects.create(
            academic_year=next_year,
            name="Term 1",
            order=1,
            start_date=date(2027, 9, 1),
            end_date=date(2027, 12, 15),
        )
        next_class = SchoolClass.objects.create(
            school=self.school,
            academic_year=next_year,
            name="JHS 1",
        )
        next_offering = SubjectOffering.objects.create(
            school=self.school,
            school_class=next_class,
            subject=self.subject,
            term=next_term,
        )
        TeacherAssignment.objects.create(
            offering=next_offering,
            teacher=self.teacher,
            is_lead=True,
        )
        self.client.force_login(self.teacher_user)

        response = self.client.get(
            reverse("gradebook_create_assessment", args=[next_offering.pk]),
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        scheme = GradeScheme.objects.get(academic_year=next_year, status=GradeScheme.Status.ACTIVE)
        self.assertEqual(scheme.categories.count(), 2)
        self.assertEqual(
            scheme.categories.aggregate(total=models.Sum("weight"))["total"],
            Decimal("100.00"),
        )
        self.assertContains(response, "Continuous assessment")
        self.assertNotContains(response, "Legacy quiz")

    def test_invalid_roster_submission_is_atomic(self):
        first, second = self.student_memberships
        self.client.force_login(self.teacher_user)
        response = self.client.post(
            reverse("gradebook_roster", args=[self.assessment.pk]),
            {f"score_{first.pk}": "18", f"score_{second.pk}": "25", "action": "publish"},
            secure=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Score cannot exceed")
        self.assertFalse(GradeEntry.objects.filter(assessment=self.assessment).exists())

    def test_teacher_can_save_draft_and_publish_roster_scores(self):
        first, second = self.student_memberships
        self.client.force_login(self.teacher_user)
        url = reverse("gradebook_roster", args=[self.assessment.pk])
        self.client.post(
            url,
            {f"score_{first.pk}": "16", f"score_{second.pk}": "14", "action": "draft"},
            secure=True,
        )
        self.assertEqual(
            set(GradeEntry.objects.filter(assessment=self.assessment).values_list("status", flat=True)),
            {GradeEntry.Status.DRAFT},
        )
        self.client.post(
            url,
            {f"score_{first.pk}": "17", f"score_{second.pk}": "15", "action": "publish"},
            secure=True,
        )
        entries = GradeEntry.objects.filter(assessment=self.assessment)
        self.assertEqual(set(entries.values_list("status", flat=True)), {GradeEntry.Status.PUBLISHED})
        self.assertEqual(set(entries.values_list("source", flat=True)), {GradeEntry.Source.MANUAL})
        self.assertEqual(set(entries.values_list("recorded_by", flat=True)), {self.teacher.pk})

    def test_roster_shows_progress_and_bulk_entry_controls(self):
        record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("17"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Ready for review",
        )
        self.client.force_login(self.teacher_user)

        response = self.client.get(
            reverse("gradebook_roster", args=[self.assessment.pk]),
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["entered_count"], 1)
        self.assertEqual(response.context["missing_count"], 1)
        self.assertEqual(response.context["pending_count"], 1)
        self.assertContains(response, "Fill blank scores with")
        self.assertContains(response, "Publish entered grades")

    def test_closed_assessment_cannot_be_changed(self):
        self.assessment.status = Assessment.Status.CLOSED
        self.assessment.save(update_fields=["status"])
        self.client.force_login(self.teacher_user)
        self.client.post(
            reverse("gradebook_roster", args=[self.assessment.pk]),
            {f"score_{self.student_memberships[0].pk}": "10", "action": "publish"},
            secure=True,
        )
        self.assertFalse(GradeEntry.objects.filter(assessment=self.assessment).exists())

    def _download_workbook(self):
        self.client.force_login(self.teacher_user)
        response = self.client.get(reverse("gradebook_template", args=[self.assessment.pk]), secure=True)
        self.assertEqual(response.status_code, 200)
        return response

    def _uploaded_workbook(self, scores, *, assessment_id=None):
        response = self._download_workbook()
        workbook = load_workbook(BytesIO(response.content))
        if assessment_id is not None:
            workbook["Instructions"]["B2"] = assessment_id
        roster = workbook["Grade Roster"]
        for row_number, score in enumerate(scores, start=2):
            roster.cell(row_number, 4).value = score
        output = BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "completed-grades.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_downloaded_template_has_bound_metadata_and_editable_score_cells(self):
        response = self._download_workbook()
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Instructions", "Grade Roster"])
        self.assertEqual(workbook["Instructions"]["B2"].value, self.assessment.pk)
        roster = workbook["Grade Roster"]
        self.assertEqual(roster.freeze_panes, "A2")
        self.assertTrue(roster.protection.sheet)
        self.assertFalse(roster["D2"].protection.locked)
        self.assertEqual(roster["A2"].value, self.student_memberships[0].pk)

    def test_unassigned_teacher_cannot_download_template(self):
        self.client.force_login(self.other_teacher_user)
        response = self.client.get(reverse("gradebook_template", args=[self.assessment.pk]), secure=True)
        self.assertEqual(response.status_code, 404)

    def test_valid_workbook_previews_then_imports_atomically(self):
        workbook = self._uploaded_workbook(["18", "16"])
        response = self.client.post(
            reverse("gradebook_import", args=[self.assessment.pk]),
            {"workbook": workbook},
            secure=True,
        )
        batch = GradeImportBatch.objects.get()
        self.assertRedirects(response, reverse("gradebook_import_preview", args=[batch.pk]), fetch_redirect_response=False)
        self.assertEqual(batch.valid_count, 2)
        self.assertEqual(batch.error_count, 0)
        self.assertFalse(GradeEntry.objects.exists())

        response = self.client.post(
            reverse("gradebook_import_confirm", args=[batch.pk]),
            {"action": "publish"},
            secure=True,
        )
        self.assertRedirects(response, reverse("gradebook_roster", args=[self.assessment.pk]), fetch_redirect_response=False)
        batch.refresh_from_db()
        self.assertEqual(batch.status, GradeImportBatch.Status.CONFIRMED)
        self.assertEqual(batch.confirmed_by, self.teacher)
        entries = GradeEntry.objects.filter(assessment=self.assessment)
        self.assertEqual(entries.count(), 2)
        self.assertEqual(set(entries.values_list("source", flat=True)), {GradeEntry.Source.IMPORT})
        self.assertEqual(set(entries.values_list("status", flat=True)), {GradeEntry.Status.PUBLISHED})
        self.assertEqual(
            set(batch.rows.values_list("status", flat=True)), {GradeImportRow.Status.IMPORTED}
        )

    def test_invalid_workbook_creates_preview_but_cannot_partially_import(self):
        workbook = self._uploaded_workbook(["18", "25"])
        self.client.post(
            reverse("gradebook_import", args=[self.assessment.pk]),
            {"workbook": workbook},
            secure=True,
        )
        batch = GradeImportBatch.objects.get()
        self.assertEqual(batch.valid_count, 1)
        self.assertEqual(batch.error_count, 1)
        response = self.client.post(
            reverse("gradebook_import_confirm", args=[batch.pk]),
            {"action": "publish"},
            secure=True,
        )
        self.assertRedirects(response, reverse("gradebook_import_preview", args=[batch.pk]), fetch_redirect_response=False)
        self.assertFalse(GradeEntry.objects.exists())
        batch.refresh_from_db()
        self.assertEqual(batch.status, GradeImportBatch.Status.PREVIEW)

    def test_workbook_for_different_assessment_is_rejected_before_batch_creation(self):
        workbook = self._uploaded_workbook(["10", "12"], assessment_id=999999)
        response = self.client.post(
            reverse("gradebook_import", args=[self.assessment.pk]),
            {"workbook": workbook},
            secure=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "different school or assessment")
        self.assertFalse(GradeImportBatch.objects.exists())

    def test_manual_grade_records_immutable_revision_history(self):
        entry, changed = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("17"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Initial teacher entry",
        )
        self.assertTrue(changed)
        self.assertEqual(entry.review_status, GradeEntry.ReviewStatus.PENDING)
        revision = entry.revisions.get()
        self.assertEqual(revision.change_type, GradeEntryRevision.ChangeType.CREATED)
        self.assertEqual(revision.new_score, Decimal("17"))
        revision.reason = "Changed history"
        with self.assertRaisesMessage(ValidationError, "immutable"):
            revision.save()

    def test_grade_change_requires_reason(self):
        with self.assertRaisesMessage(ValidationError, "reason is required"):
            record_grade_entry(
                school=self.school,
                assessment=self.assessment,
                student=self.student_memberships[0],
                actor=self.teacher,
                score=Decimal("17"),
                source=GradeEntry.Source.MANUAL,
                status=GradeEntry.Status.DRAFT,
                reason="",
            )

    def test_admin_review_locks_and_return_unlocks_grade_correction(self):
        entry, _ = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("17"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Ready for review",
        )
        admin_user = User.objects.create_user(username="school-admin", password="test-password")
        administrator = SchoolMembership.objects.create(
            school=self.school, user=admin_user, role=SchoolMembership.Role.SCHOOL_ADMIN
        )
        review_grade_entry(
            entry=entry,
            reviewer=administrator,
            decision=GradeReviewDecision.Decision.APPROVED,
            note="Checked against source sheet.",
        )
        with self.assertRaisesMessage(ValidationError, "must be returned"):
            record_grade_entry(
                school=self.school,
                assessment=self.assessment,
                student=self.student_memberships[0],
                actor=self.teacher,
                score=Decimal("18"),
                source=GradeEntry.Source.MANUAL,
                status=GradeEntry.Status.PUBLISHED,
                reason="Teacher correction",
            )
        review_grade_entry(
            entry=entry,
            reviewer=administrator,
            decision=GradeReviewDecision.Decision.RETURNED,
            note="Please correct the source total.",
        )
        corrected, changed = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("18"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Corrected source total",
        )
        self.assertTrue(changed)
        self.assertEqual(corrected.review_status, GradeEntry.ReviewStatus.PENDING)
        self.assertEqual(corrected.revisions.count(), 2)
        self.assertEqual(corrected.review_decisions.count(), 2)

    def test_teacher_cannot_review_grade(self):
        entry, _ = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("12"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Teacher entry",
        )
        with self.assertRaises(PermissionDenied):
            review_grade_entry(
                entry=entry,
                reviewer=self.teacher,
                decision=GradeReviewDecision.Decision.APPROVED,
            )

    def test_teacher_cannot_open_administrator_review_queue(self):
        self.client.force_login(self.teacher_user)
        response = self.client.get(reverse("gradebook_review_queue"), secure=True)
        self.assertRedirects(response, reverse("home"), fetch_redirect_response=False)

    def test_administrator_can_approve_grade_through_review_queue(self):
        entry, _ = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("13"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Submit for administrator review",
        )
        admin_user = User.objects.create_user(username="review-admin", password="test-password")
        administrator = SchoolMembership.objects.create(
            school=self.school, user=admin_user, role=SchoolMembership.Role.SCHOOL_ADMIN
        )
        self.client.force_login(admin_user)
        response = self.client.post(
            reverse("gradebook_review", args=[entry.pk]),
            {"decision": GradeReviewDecision.Decision.APPROVED, "note": "Verified"},
            secure=True,
        )
        self.assertRedirects(response, reverse("gradebook_review_queue"), fetch_redirect_response=False)
        entry.refresh_from_db()
        self.assertEqual(entry.review_status, GradeEntry.ReviewStatus.APPROVED)
        self.assertEqual(entry.reviewed_by, administrator)

        queue = self.client.get(reverse("gradebook_review_queue"), secure=True)
        self.assertContains(queue, "Approved and locked")
        self.assertNotContains(queue, f'action="{reverse("gradebook_review", args=[entry.pk])}"')
        with self.assertRaisesMessage(ValidationError, "already approved"):
            review_grade_entry(
                entry=entry,
                reviewer=administrator,
                decision=GradeReviewDecision.Decision.APPROVED,
            )

    def test_review_queue_prioritizes_pending_and_filters_status(self):
        pending, _ = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("14"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Pending score",
        )
        approved, _ = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[1],
            actor=self.teacher,
            score=Decimal("16"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.PUBLISHED,
            reason="Approved score",
        )
        admin_user = User.objects.create_user(username="filter-admin", password="test-password")
        administrator = SchoolMembership.objects.create(
            school=self.school,
            user=admin_user,
            role=SchoolMembership.Role.SCHOOL_ADMIN,
        )
        review_grade_entry(
            entry=approved,
            reviewer=administrator,
            decision=GradeReviewDecision.Decision.APPROVED,
            note="Verified",
        )
        self.client.force_login(admin_user)

        queue = self.client.get(reverse("gradebook_review_queue"), secure=True)

        self.assertEqual(queue.status_code, 200)
        entries = list(queue.context["entries"])
        self.assertEqual(entries[0], pending)
        self.assertEqual(queue.context["counts"]["pending"], 1)
        self.assertEqual(queue.context["counts"]["approved"], 1)

        approved_only = self.client.get(
            reverse("gradebook_review_queue"),
            {"status": GradeEntry.ReviewStatus.APPROVED},
            secure=True,
        )
        self.assertEqual(list(approved_only.context["entries"]), [approved])

    def test_unassigned_teacher_cannot_open_grade_correction(self):
        entry, _ = record_grade_entry(
            school=self.school,
            assessment=self.assessment,
            student=self.student_memberships[0],
            actor=self.teacher,
            score=Decimal("11"),
            source=GradeEntry.Source.MANUAL,
            status=GradeEntry.Status.DRAFT,
            reason="Initial draft",
        )
        self.client.force_login(self.other_teacher_user)
        response = self.client.get(reverse("gradebook_correct", args=[entry.pk]), secure=True)
        self.assertEqual(response.status_code, 404)

    def test_quiz_sync_averages_attempts_and_is_idempotent(self):
        quiz = Quiz.objects.create(
            subject=self.subject, teacher=self.teacher_user, title="Legacy quiz"
        )
        first, second = self.student_memberships
        Submission.objects.create(quiz=quiz, student=first.user, score=80)
        Submission.objects.create(quiz=quiz, student=first.user, score=100)
        Submission.objects.create(quiz=quiz, student=second.user, score=60)
        self.assessment.legacy_quiz = quiz
        self.assessment.save(update_fields=["legacy_quiz"])
        first_result = sync_legacy_assessment(assessment=self.assessment, actor=self.teacher)
        second_result = sync_legacy_assessment(assessment=self.assessment, actor=self.teacher)
        self.assertEqual(first_result, {"source": "quiz 'Legacy quiz'", "found": 2, "changed": 2, "unchanged": 0})
        self.assertEqual(second_result["changed"], 0)
        self.assertEqual(second_result["unchanged"], 2)
        self.assertEqual(
            GradeEntry.objects.get(assessment=self.assessment, student=first).score,
            Decimal("18.00"),
        )
        self.assertEqual(GradeEntryRevision.objects.filter(entry__assessment=self.assessment).count(), 2)

    def test_assignment_sync_uses_latest_finalized_submission(self):
        assignment = Assignment.objects.create(
            subject=self.subject, teacher=self.teacher_user, title="Legacy assignment"
        )
        student = self.student_memberships[0]
        AssignmentSubmission.objects.create(
            assignment=assignment,
            student=student.user,
            file="assignment_submissions/first.pdf",
            final_score=50,
            max_score=100,
        )
        AssignmentSubmission.objects.create(
            assignment=assignment,
            student=student.user,
            file="assignment_submissions/latest.pdf",
            final_score=80,
            max_score=100,
        )
        self.assessment.legacy_assignment = assignment
        self.assessment.save(update_fields=["legacy_assignment"])
        result = sync_legacy_assessment(assessment=self.assessment, actor=self.teacher)
        self.assertEqual(result["changed"], 1)
        self.assertEqual(
            GradeEntry.objects.get(assessment=self.assessment, student=student).score,
            Decimal("16.00"),
        )
